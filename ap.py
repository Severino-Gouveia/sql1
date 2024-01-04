import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

# Configurações de conexão com o banco de dados
SERVIDOR = '172.20.20.38'
BANCO_DE_DADOS = 'EAMPROD'
USUARIO = 'excel'
SENHA = 'excel'
NOME_DO_DRIVER = 'ODBC Driver 17 for SQL Server'

# String de conexão formatada
string_de_conexao = f'DRIVER={{{NOME_DO_DRIVER}}};SERVER={SERVIDOR};DATABASE={BANCO_DE_DADOS};UID={USUARIO};PWD={SENHA}'

# Mensagem indicando o início da tentativa de conexão
print("Conectando ao banco de dados...")

try:
    # Tentativa de conexão com o bloco "with" para garantir o fechamento correto
    with pyodbc.connect(string_de_conexao) as conexao:
        # Mensagem indicando uma conexão bem-sucedida
        print("Conectado ao banco de dados com sucesso.")

        # Consulta SQL para selecionar os primeiros 5 registros
        consulta_sql = '''
            SELECT R5ADDETAILS.ADD_CREATED as DATA_TEXTO, R5ADDETAILS.ADD_CODE as AES,
            R5EVENTS.EVT_DESC as Descricao, CASE R5EVENTS.EVT_STATUS
                WHEN 'FE' THEN 'Fechado'
                WHEN 'RP' THEN 'Reprogramar'
                WHEN 'CM' THEN 'Aguardando Chegada Material'
                WHEN 'P' THEN 'Programada'
                WHEN 'REJ' THEN 'Rejeitada'
                WHEN 'AP' THEN 'Aprovada'
                WHEN 'PM' THEN 'Aguardar Parada de Máquina'
                WHEN 'R' THEN 'Emitido'
                WHEN 'CL' THEN 'Concluída'
                WHEN 'EE' THEN 'Em Execução'
                ELSE 'OUTROS'
            END as Stat_AES,
            CASE R5ADDETAILS.ADD_TYPE
                WHEN '*' THEN 'Sim'
                ELSE 'Não'
            END as Observacao,
            R5EVENTS.EVT_STATUS as EVT_STAT, R5ADDETAILS.ADD_LINE as Linha,
            R5ADDETAILS.ADD_TEXT as texto, R5ADDETAILS.ADD_USER as matricula,
            R5PERSONNEL.PER_DESC as Colaborador, R5CREWS.CRW_DESC AS turno
            FROM EAMPROD.dbo.R5ADDETAILS R5ADDETAILS
            INNER JOIN R5EVENTS ON R5EVENTS.EVT_CODE = R5ADDETAILS.ADD_CODE
            INNER JOIN R5PERSONNEL ON R5PERSONNEL.PER_CODE = R5ADDETAILS.ADD_USER
            INNER JOIN R5CREWEMPLOYEES ON R5CREWEMPLOYEES.CRE_PERSON = R5ADDETAILS.ADD_USER
            INNER JOIN R5CREWS ON R5CREWEMPLOYEES.CRE_CREW = R5CREWS.CRW_CODE
            WHERE R5ADDETAILS.ADD_ENTITY = 'EVNT' AND (R5ADDETAILS.ADD_CREATED >= '2023-12-06') AND (R5ADDETAILS.ADD_CREATED < '2023-12-07') AND (R5EVENTS.EVT_MRC ='DP02')
        '''

        # Bloco "with" para o cursor, que será fechado automaticamente
        with conexao.cursor() as cursor:
            # Execução da consulta SQL
            cursor.execute(consulta_sql)
            # Recuperação de todos os registros
            registros = cursor.fetchall()

            # Converter os resultados para um DataFrame do pandas
            relatorio_estilizado = pd.DataFrame.from_records(registros, columns=[desc[0] for desc in cursor.description])

            # Renomear a coluna 'Stat_AES' para 'Status'
            relatorio_estilizado.rename(columns={'Stat_AES': 'Status'}, inplace=True)

            # Extrair a informação de hora da coluna 'DATA_TEXTO'
            relatorio_estilizado['Hora'] = relatorio_estilizado['DATA_TEXTO'].dt.strftime('%H:%M:%S')

            # Selecionar apenas as colunas desejadas para o relatório estilizado
            relatorio_estilizado = relatorio_estilizado[['AES', 'Descricao', 'Colaborador', 'texto', 'Hora', 'Status']]

            # Verificar se o arquivo com o nome atual já existe
            contagem = 1
            while os.path.exists(f'RDM_{contagem}.xlsx'):
                contagem += 1

            # Criar o nome do arquivo com a contagem
            excel_file = f'RDM_{contagem}.xlsx'

            # Carregar o arquivo Excel para adicionar estilos
            wb = Workbook()
            ws = wb.active

            # Adicionar a descrição com margens
            descricao = "\nRelatório Diário de Manutenção\nDepartamento de Manutenção Elétrica\nSemana: 50 Data: 14/12/2023"
            ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=len(relatorio_estilizado.columns))
            cell = ws.cell(row=1, column=1, value=descricao)
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))

       

            # Adicionar cabeçalho
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            border = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))

            # Adicionar cabeçalho
            for col_num, value in enumerate(relatorio_estilizado.columns, 1):
                cell = ws.cell(row=5, column=col_num, value=value)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Adicionar cores alternadas às linhas
            for r_idx, row in enumerate(relatorio_estilizado.iterrows(), start=6):
                for c_idx, value in enumerate(row[1], start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                    # Adicionar cores alternadas
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

            # Adicionar estilo condicional às células na coluna 'Status'
            for r_idx, row in enumerate(relatorio_estilizado.iterrows(), start=6):
                for c_idx, value in enumerate(row[1], start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                    # Adicionar estilo condicional à coluna 'Status'
                    if c_idx == relatorio_estilizado.columns.get_loc('Status') + 1:
                        if value == 'Fechado':
                            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        elif value == 'Reprogramar':
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        # Adicione mais condições conforme necessário

            # Adicionar estilo adicional às células
            for col_num, value in enumerate(relatorio_estilizado.columns, 1):
                cell = ws.cell(row=5, column=col_num, value=value)
                cell.font = Font(bold=True, color="FFFFFF", size=12)  # Ajuste o tamanho da fonte conforme necessário
                cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Ajustar largura das colunas
            for col_num, column in enumerate(ws.columns, 1):
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                col_letter = get_column_letter(col_num)  # Obter a letra da coluna
                ws.column_dimensions[col_letter].width = adjusted_width

            # Salvar o arquivo estilizado
            wb.save(excel_file)

            # Mensagem indicando o término da impressão
            print(f"Conclusão da criação {excel_file}.")

except pyodbc.Error as erro:
    # Em caso de erro durante a conexão ou execução da consulta, imprimir a mensagem de erro
    print(f"Erro: {erro}")






